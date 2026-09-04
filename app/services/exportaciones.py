import json
import os
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.models.equipo import Equipo
from app.models.orden import OrdenServicio

TIPOS_EXPORTACION = {"ordenes": "Órdenes de servicio", "clientes": "Clientes", "diagnosticos": "Diagnósticos", "completo": "Reporte completo"}


def _valor(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, datetime):
        return valor.isoformat(sep=" ", timespec="seconds")
    if isinstance(valor, date):
        return valor.isoformat()
    return valor if valor is not None else ""


def obtener_filas(db: Session, tipo: str, fecha_inicio: date | None = None, fecha_fin: date | None = None, estado: str | None = None, tecnico: str | None = None):
    consulta = db.query(OrdenServicio).options(joinedload(OrdenServicio.equipo).joinedload(Equipo.cliente), joinedload(OrdenServicio.diagnostico))
    if fecha_inicio:
        consulta = consulta.filter(OrdenServicio.fecha_ingreso >= datetime.combine(fecha_inicio, time.min))
    if fecha_fin:
        consulta = consulta.filter(OrdenServicio.fecha_ingreso <= datetime.combine(fecha_fin, time.max))
    if estado:
        consulta = consulta.filter(OrdenServicio.estado == estado)
    if tecnico:
        consulta = consulta.filter(OrdenServicio.tecnico_responsable == tecnico)
    filas = []
    for orden in consulta.order_by(OrdenServicio.id).all():
        equipo, cliente, diagnostico = orden.equipo, orden.equipo.cliente, orden.diagnostico
        base = {"id_orden": orden.id, "numero_orden": orden.numero_orden, "fecha_ingreso": orden.fecha_ingreso, "estado": orden.estado, "tecnico_responsable": orden.tecnico_responsable}
        if tipo == "ordenes":
            fila = {**base, "falla_reportada": orden.falla_reportada}
        elif tipo == "clientes":
            fila = {**base, "id_cliente": cliente.id, "nombres": cliente.nombres, "apellidos": cliente.apellidos, "dni_ruc": cliente.dni_ruc, "telefono": cliente.telefono}
        elif tipo == "diagnosticos":
            fila = {**base, "id_diagnostico": diagnostico.id if diagnostico else None, "falla_encontrada": diagnostico.falla_encontrada if diagnostico else None, "solucion_recomendada": diagnostico.solucion_recomendada if diagnostico else None, "repuestos_necesarios": diagnostico.repuestos_necesarios if diagnostico else None, "costo_estimado": diagnostico.costo_estimado if diagnostico else None, "fecha_diagnostico": diagnostico.fecha_diagnostico if diagnostico else None}
        else:
            fila = {**base, "falla_reportada": orden.falla_reportada, "id_cliente": cliente.id, "cliente": f"{cliente.nombres} {cliente.apellidos}", "dni_ruc": cliente.dni_ruc, "telefono": cliente.telefono, "id_equipo": equipo.id, "tipo_equipo": equipo.tipo, "marca": equipo.marca, "modelo": equipo.modelo, "numero_serie": equipo.numero_serie, "falla_encontrada": diagnostico.falla_encontrada if diagnostico else None, "solucion_recomendada": diagnostico.solucion_recomendada if diagnostico else None, "costo_estimado": diagnostico.costo_estimado if diagnostico else None}
        filas.append({clave: _valor(valor) for clave, valor in fila.items()})
    return filas


def crear_excel(filas: list[dict], titulo: str) -> BytesIO:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Exportación"
    encabezados = list(filas[0].keys()) if filas else ["id_orden"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="C1121F")
    for fila in filas:
        hoja.append([fila.get(encabezado, "") for encabezado in encabezados])
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    for columna in hoja.columns:
        hoja.column_dimensions[columna[0].column_letter].width = min(max(len(str(celda.value or "")) for celda in columna) + 2, 45)
    libro.properties.title = titulo
    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    return salida


def sincronizar_google_sheets(filas: list[dict]) -> tuple[int, int]:
    credenciales = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    spreadsheet_id = os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID")
    nombre_hoja = os.getenv("GOOGLE_SHEETS_WORKSHEET", "Exportaciones")
    if not credenciales or not spreadsheet_id:
        raise RuntimeError("Configura GOOGLE_SERVICE_ACCOUNT_JSON y GOOGLE_SHEETS_SPREADSHEET_ID en .env.")
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError("Instala las dependencias de Google indicadas en requirements.txt.") from exc
    info = json.loads(credenciales) if credenciales.lstrip().startswith("{") else None
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    cuenta = Credentials.from_service_account_info(info, scopes=scopes) if info else Credentials.from_service_account_file(credenciales, scopes=scopes)
    hojas = build("sheets", "v4", credentials=cuenta, cache_discovery=False).spreadsheets()
    existente = hojas.values().get(spreadsheetId=spreadsheet_id, range=f"'{nombre_hoja}'!A:ZZ").execute().get("values", [])
    encabezados = list(filas[0].keys()) if filas else ["id_orden"]
    if not existente:
        hojas.values().update(spreadsheetId=spreadsheet_id, range=f"'{nombre_hoja}'!A1", valueInputOption="RAW", body={"values": [encabezados]}).execute()
    por_id = {str(fila[0]): indice for indice, fila in enumerate(existente[1:], start=2) if fila}
    creados = actualizados = 0
    for fila in filas:
        valores, clave = [[fila.get(campo, "") for campo in encabezados]], str(fila["id_orden"])
        if clave in por_id:
            hojas.values().update(spreadsheetId=spreadsheet_id, range=f"'{nombre_hoja}'!A{por_id[clave]}", valueInputOption="RAW", body={"values": valores}).execute()
            actualizados += 1
        else:
            hojas.values().append(spreadsheetId=spreadsheet_id, range=f"'{nombre_hoja}'!A:ZZ", valueInputOption="RAW", insertDataOption="INSERT_ROWS", body={"values": valores}).execute()
            creados += 1
    return creados, actualizados
